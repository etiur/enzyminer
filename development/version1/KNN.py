from sklearn.model_selection import train_test_split
from sklearn.neighbors import KNeighborsClassifier as KNN
from sklearn.neighbors import NeighborhoodComponentsAnalysis as NCA
from sklearn.pipeline import Pipeline
import pandas as pd
from sklearn.metrics import matthews_corrcoef, confusion_matrix
from sklearn.metrics import classification_report as class_re
from sklearn.model_selection import GridSearchCV
from os import path
from collections import namedtuple
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
from openpyxl import Workbook


"""classification functions"""
def knn_classification_nca(X_train, Y_train, X_test, state=20):
    """A function that applies grid and random search to tune model and also gives a prediction, it also uses the
    NCA transformation of the data which seems to improve performance"""

    # Creating a score and parameters to search from
    scoring = {"f1": "f1_weighted"}

    grid_param2_nca = { "knn__n_neighbors": range(1, 11), "knn__p": range(1, 6), "knn__metric": ["minkowski", "canberra", "hamming"]}

    with_nca = namedtuple("with_nca", ["fitted_grid", "y_grid", "grid_train_Y"])

    # Model setting
    nca = NCA(random_state=state)
    pipe = Pipeline(steps=[("nca", nca), ("knn", KNN(n_jobs=-1))])
    knn_grid = GridSearchCV(pipe, grid_param2_nca, scoring=scoring, refit="f1", cv=5)

    # Model training with nca
    fitted_grid = knn_grid.fit(X_train, Y_train)
    # Model predictions with nca
    y_grid = fitted_grid.best_estimator_.predict(X_test)
    # training data prediction with nca
    grid_train_Y = fitted_grid.best_estimator_.predict(X_train)
    nca_model_list = with_nca(*[fitted_grid, y_grid, grid_train_Y])

    return nca_model_list

def knn_classification(X_train, Y_train, X_test):
    """A function that applies grid and random search to tune model and also gives a prediction, it also uses the
    NCA transformation of the data which seems to improve performance"""

    # Creating a score and parameters to search from
    scoring = {"f1": "f1_weighted"}
    grid_param2 = {"n_neighbors": range(1, 11), "p": range(1,6), "metric": ["minkowski", "canberra", "hamming"]}

    no_nca = namedtuple("no_nca", ["fitted_grid", "y_grid", "grid_train_Y"])

    # Model setting
    knn_grid = GridSearchCV(KNN(n_jobs=-1), grid_param2, scoring=scoring, refit="f1", cv=5)

    # Model training
    fitted_grid = knn_grid.fit(X_train, Y_train)
    # Model predictions
    y_grid = fitted_grid.best_estimator_.predict(X_test)
    # training data prediction
    grid_train_Y = fitted_grid.best_estimator_.predict(X_train)

    no_nca_model_list = no_nca(*[fitted_grid, y_grid, grid_train_Y])

    return no_nca_model_list

def print_score(no_nca_list, Y_train, Y_test):
    """ The function prints the scores of the models and the prediction performance """

    target_names = ["class 0", "class 1"]
    no_nca = namedtuple("no_nca", ["grid_score", "grid_params", "grid_confusion", "tr_report", "te_report",
                                    "train_mat", "grid_matthews", "grid_train_confusion"])
    # Model comparison
    grid_score = no_nca_list.fitted_grid.best_score_
    grid_params = no_nca_list.fitted_grid.best_params_

    # Training scores
    grid_train_confusion = confusion_matrix(Y_train, no_nca_list.grid_train_Y)
    grid_tr_report = class_re(Y_train, no_nca_list.grid_train_Y, target_names=target_names, output_dict=True)
    train_mat = matthews_corrcoef(Y_train, no_nca_list.grid_train_Y)

    # Test metrics grid
    grid_confusion = confusion_matrix(Y_test, no_nca_list.y_grid)
    grid_matthews = matthews_corrcoef(Y_test, no_nca_list.y_grid)
    grid_te_report = class_re(Y_test, no_nca_list.y_grid, target_names=target_names, output_dict=True)

    everything = no_nca(*[grid_score, grid_params, grid_confusion, grid_tr_report, grid_te_report, train_mat,
           grid_matthews, grid_train_confusion])

    return everything

def nested_cv(X, Y):
    """Performs something similar to a nested cross-validation"""
    metric_list_nca = []
    metric_list_no_nca = []
    model_list = []
    parameter_list_nca = []
    parameter_list_no_nca =[]
    random_state = [20, 40, 70, 80, 90]
    scaling = MinMaxScaler()
    esterase = ['EH51(22)', 'EH75(16)', 'EH46(23)', 'EH98(11)', 'EH49(23)']

    for states in random_state:
        X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.20, random_state=states, stratify=Y)

        X_train = X_train.loc[[x for x in X_train.index if x not in esterase]]
        X_test = X_test.loc[[x for x in X_test.index if x not in esterase]]
        Y_train = Y_train.loc[[x for x in Y_train.index if x not in esterase]]
        Y_test = Y_test.loc[[x for x in Y_test.index if x not in esterase]]


        transformed_x = scaling.fit_transform(X_train)
        transformed_x = pd.DataFrame(transformed_x)
        transformed_x.index = X_train.index
        transformed_x.columns = X_train.columns

        test_x = scaling.transform(X_test)
        test_x = pd.DataFrame(test_x)
        test_x.index = X_test.index
        test_x.columns = X_test.columns

        # nca classfication
        nca_model_list = knn_classification_nca(transformed_x, Y_train, test_x)
        nca_score = print_score(nca_model_list, Y_train, Y_test)
        metric_list_nca.append([nca_score.grid_score, nca_score.train_mat, nca_score.grid_matthews])
        parameter_list_nca.append(
            [nca_score.grid_params, nca_score.grid_confusion, nca_score.tr_report, nca_score.te_report,
             nca_score.grid_train_confusion])

        # no nca classification
        no_nca_model_list = knn_classification(transformed_x, Y_train, test_x)
        no_nca_score = print_score(no_nca_model_list, Y_train, Y_test)
        metric_list_no_nca.append([no_nca_score.grid_score, no_nca_score.train_mat, no_nca_score.grid_matthews])
        parameter_list_no_nca.append(
            [no_nca_score.grid_params, no_nca_score.grid_confusion, no_nca_score.tr_report, no_nca_score.te_report,
             no_nca_score.grid_train_confusion])

        # puts the models into a list
        model_list.append([nca_model_list.fitted_grid, nca_model_list.y_grid, no_nca_model_list.fitted_grid, no_nca_model_list.y_grid])

    return model_list, metric_list_nca, metric_list_no_nca, parameter_list_nca, parameter_list_no_nca, random_state

def mean_nested(X, Y):
    """From the results of the nested_CV it computes the means of the different performance metrics"""
    model_list, metric_list_nca, metric_list_no_nca, parameter_list_nca, parameter_list_no_nca, random_state = nested_cv(X, Y)

    score_record = namedtuple("scores", ["grid_score", "train_mat", "grid_matthews"])

    parameter_record = namedtuple("parameters", ["grid_params", "grid_confusion", "tr_report", "te_report",
                                                 "grid_train_confusion"])

    model_record = namedtuple("models", ["nca_fitted", "nca_y", "no_nca_fitted", "no_nca_y"])

    # with nca
    parameters_nca = [parameter_record(*z) for z in parameter_list_nca]
    records_nca = [score_record(*y) for y in metric_list_nca]
    named_models = [model_record(*d) for d in model_list]

    # Without nca
    parameters_no_nca = [parameter_record(*z) for z in parameter_list_no_nca]
    records_no_nca = [score_record(*y) for y in metric_list_no_nca]


    return named_models, parameters_nca, records_nca, parameters_no_nca, records_no_nca, random_state

def unlisting(parameters, records, mode=1):
    """ A function that separates all the scores in independent lists"""
    # Getting all scores grid search
    g_mathew = [x.grid_matthews for x in records]
    train_mat = [x.train_mat for x in records]
    cv_score = [x.grid_score for x in records]

    if mode == 1:
        g_neighbours = [y.grid_params["n_neighbors"] for y in parameters]
        g_p = [y.grid_params["p"] for y in parameters]
        g_distance = [y.grid_params["metric"] for y in parameters]

    else:
        g_neighbours = [y.grid_params["knn__n_neighbors"] for y in parameters]
        g_p = [y.grid_params["knn__p"] for y in parameters]
        g_distance = [y.grid_params["knn__metric"] for y in parameters]

    return g_mathew, train_mat, cv_score, g_neighbours, g_p, g_distance

def to_dataframe(parameters, records, random_state, mode):
    matrix = namedtuple("confusion_matrix", ["true_n", "false_p", "false_n", "true_p"])

    g_mathew, train_mat, cv_score, g_neighbours, g_p, g_distance = unlisting(parameters, records, mode)

    # Taking the confusion matrix
    g_test_confusion = [matrix(*x.grid_confusion.ravel()) for x in parameters]
    g_training_confusion = [matrix(*x.grid_train_confusion.ravel()) for x in parameters]
    g_te_report = [pd.DataFrame(x.te_report).transpose() for x in parameters]
    g_tr_report = [pd.DataFrame(x.tr_report).transpose() for x in parameters]

    # Separating confusion matrix into individual elements
    g_test_true_n = [y.true_n for y in g_test_confusion]
    g_test_false_p = [y.false_p for y in g_test_confusion]
    g_test_false_n = [y.false_n for y in g_test_confusion]
    g_test_true_p = [y.true_p for y in g_test_confusion]

    g_training_true_n = [z.true_n for z in g_training_confusion]
    g_training_false_p = [z.false_p for z in g_training_confusion]
    g_training_false_n = [z.false_n for z in g_training_confusion]
    g_training_true_p = [z.true_p for z in g_training_confusion]


    g_dataframe = pd.DataFrame([random_state, g_neighbours, g_p, g_distance, g_test_true_n, g_test_true_p,
                g_test_false_p, g_test_false_n, g_training_true_n, g_training_true_p, g_training_false_p,
                g_training_false_n, g_mathew, train_mat])

    g_dataframe = g_dataframe.transpose()
    g_dataframe.columns = ["random state","neighbours", "p", "distance", "test_tn", "test_tp", "test_fp", "test_fn",
                           "train_tn", "train_tp", "train_fp", "train_fn", "Matthews", "train_Mat"]

    return g_dataframe, g_te_report, g_tr_report

def writing(dataset1, nca_te_report, nca_tr_report, dataset2, te_report, tr_report, sheet_name, row=0):

    if not path.exists(f"knn_no_error.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename=f"knn_no_error.xlsx")

    book = load_workbook(f"knn_no_error.xlsx")
    writer = pd.ExcelWriter(f"knn_no_error.xlsx", engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    dataset1.to_excel(writer, sheet_name=f'{sheet_name}', startrow=row)
    nca_train = pd.concat(nca_tr_report, axis=1)
    nca_test = pd.concat(nca_te_report, axis=1)
    nca_test.to_excel(writer, sheet_name=f'{sheet_name}', startrow=row, startcol=len(dataset1.columns) + 3)
    nca_train.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(nca_test.index)+3,
                           startcol=len(dataset1.columns) + 3)

    dataset2.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(dataset1.index) + 3)
    no_nca_train = pd.concat(tr_report, axis=1)
    no_nca_test = pd.concat(te_report, axis=1)
    no_nca_test.to_excel(writer, sheet_name=f'{sheet_name}', startrow=(len(nca_test.index) + 3)*2, startcol=len(dataset2.columns)+3)
    no_nca_train.to_excel(writer, sheet_name=f'{sheet_name}', startrow=(len(nca_test.index) + 3) * 3,
                     startcol=len(dataset2.columns) + 3)

    writer.save()
    writer.close()

def run(sheet):
    """A function that performs classification given an excel sheet's name"""
    VHSE = pd.read_excel("sequences.xlsx", index_col=0, sheet_name="VHSE", engine='openpyxl')
    Y = VHSE["label1"].copy()
    X = pd.read_excel(f"esterase_noerror.xlsx", index_col=0, sheet_name=f"{sheet}", engine='openpyxl')
    if X.isnull().values.any():
        X.dropna(axis=1, inplace=True)
        X.drop(["go"], axis=1, inplace=True)
    # to hold the data generated
    results = namedtuple("results",
                         ["named_models", "parameters_nca", "records_nca", "parameters_no_nca", "records_no_nca"])
    data = namedtuple("dataframe", ["dataframe", "te_report", "tr_report"])

    named_models, parameters_nca, records_nca, parameters_no_nca, records_no_nca, random_state = mean_nested(X, Y)
    nca_dataframe, nca_te_report, nca_tr_report = to_dataframe(parameters_nca, records_nca, random_state, mode=0)
    no_nca_dataframe, no_nca_te_report, no_nca_tr_report = to_dataframe(parameters_no_nca, records_no_nca, random_state, mode=1)

    results_list = results(*[named_models, parameters_nca, records_nca, parameters_no_nca, records_no_nca])
    data_nca = data(*[nca_dataframe, nca_te_report, nca_tr_report])
    data_no_nca = data(*[no_nca_dataframe, no_nca_te_report, no_nca_tr_report])

    return results_list, data_nca, data_no_nca

def run_all():

    book = load_workbook(f"esterase_noerror.xlsx")
    result_list = []
    for ws in book.worksheets:
        if ws.title == "random_30" or ws.title == "ch2_20":
            results, data_nca, data_no_nca = run(ws.title)
            writing(data_nca.dataframe, data_nca.te_report, data_nca.tr_report, data_no_nca.dataframe, data_no_nca.te_report,
                data_no_nca.tr_report, ws.title)
            result_list.append(results)

    return result_list

run_all()