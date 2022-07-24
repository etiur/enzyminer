from sklearn.neighbors import KNeighborsClassifier as KNN
from sklearn.neighbors import NeighborhoodComponentsAnalysis as NCA
from sklearn.pipeline import Pipeline
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
from utils_generate import split_transform, print_score, to_dataframe, vote
from os import path


def fit(state=1):
    """A function that trains the classifiers and ensembles them"""
    # the features and the labels
    VHSE = pd.read_excel("sequences.xlsx", index_col=0, engine='openpyxl')
    Y = VHSE["label"].copy()
    X_knn_random = pd.read_excel("esterase_binary.xlsx", index_col=0, sheet_name="random_30", engine='openpyxl')
    X_knn_ch = pd.read_excel("esterase_binary.xlsx", index_col=0, sheet_name="ch2_20", engine='openpyxl')

    if X_knn_random.isnull().values.any():
        X_knn_random.dropna(axis=1, inplace=True)
        X_knn_random.drop(["go"], axis=1, inplace=True)
    if X_knn_ch.isnull().values.any():
        X_knn_ch.dropna(axis=1, inplace=True)
        X_knn_ch.drop(["go"], axis=1, inplace=True)

    if state == 1:
        model_dict = {20: KNN(n_neighbors=7, p=5, metric="minkowski", n_jobs=-1),
                  40: KNN(n_neighbors=5, p=5, metric="minkowski", n_jobs=-1),
                  70: KNN(n_neighbors=7, p=1, metric="minkowski", n_jobs=-1),
                  80: KNN(n_neighbors=4, p=4, metric="minkowski", n_jobs=-1),
                  90: KNN(n_neighbors=9, p=1, metric="minkowski", n_jobs=-1)}
        for states in [20, 40, 70, 80, 90]:
            # split and train
            transformed_x_knn, test_x_knn, Y_train_knn, Y_test_knn, X_train_knn, X_test_knn = split_transform(
                X_knn_random, Y,
                states)
            model_dict[states].fit(transformed_x_knn, Y_train_knn)

        for states in [20, 40, 70, 80, 90]:
        # split and train
            transformed_x_knn, test_x_knn, Y_train_knn, Y_test_knn, X_train_knn, X_test_knn = split_transform(X_knn_ch, Y,
                                                                                                          states)
            model_dict[states].fit(transformed_x_knn, Y_train_knn)

    return model_dict


def predict(fitted, test_x, transformed_x):
    """Using fitted models it makes predictions"""
    preditions_train = {}
    predictions_test = {}
    for states in [20, 40, 70, 80, 90]:
        # predict on X_test
        predictions_test[states] = fitted[states].predict(test_x)
        # predict on X_train
        preditions_train[states] = fitted[states].predict(transformed_x)

    return predictions_test, preditions_train


def get_scores(Y_test_svc, Y_train_svc, test_pred, train_pred):
    """Converts the scores into dataframes"""
    # ensembles the predictions
    vote_2_test, index2_test = vote(test_pred[20], test_pred[40], test_pred[70])
    vote_2_train, index2_train = vote(train_pred[20], train_pred[40], train_pred[70])
    vote_1_test, index1_test = vote(test_pred[20], test_pred[40])
    vote_1_train, index1_train = vote(train_pred[20], train_pred[40])
    vote_3_test, index3_test = vote(test_pred[20], test_pred[40], test_pred[70], test_pred[80])
    vote_3_train, index3_train = vote(train_pred[20], train_pred[40], train_pred[70], train_pred[80])
    vote_4_test, index4_test = vote(test_pred[20], test_pred[40], test_pred[70], test_pred[80], test_pred[90])
    vote_4_train, index4_train = vote(train_pred[20], train_pred[40], train_pred[70], train_pred[80], train_pred[90])

    # generating the scores
    individual_scores = []
    for states in [20, 40, 70, 80, 90]:
        score = print_score(Y_test_svc, test_pred[states], train_pred[states], Y_train_svc)
        individual_scores.append(score)

    ensemble2_purged = print_score(Y_test_svc, vote_2_test, vote_2_train, Y_train_svc, index2_test, index2_train,
                                   mode=1)
    ensemble1_purged = print_score(Y_test_svc, vote_1_test, vote_1_train, Y_train_svc, index1_test, index1_train,
                                   mode=1)
    ensemble3_purged = print_score(Y_test_svc, vote_3_test, vote_3_train, Y_train_svc, index3_test, index3_train,
                                   mode=1)
    ensemble4_purged = print_score(Y_test_svc, vote_4_test, vote_4_train, Y_train_svc, index4_test, index4_train,
                                   mode=1)

    # put all the sores into dataframe
    dataframe_20, te_report_20, tr_report_20 = to_dataframe(individual_scores[0], ["knn_20"])
    dataframe_40, te_report_40, tr_report_40 = to_dataframe(individual_scores[1], ["knn_40"])
    dataframe_70, te_report_70, tr_report_70 = to_dataframe(individual_scores[2], ["knn_70"])
    dataframe_80, te_report_80, tr_report_80 = to_dataframe(individual_scores[3], ["knn_80"])
    dataframe_90, te_report_90, tr_report_90 = to_dataframe(individual_scores[4], ["knn_90"])
    dataframe_ense1, te_report_ense1, tr_report_ense1 = to_dataframe(ensemble1_purged, ["ensemble1"])
    dataframe_ense2, te_report_ense2, tr_report_ense2 = to_dataframe(ensemble2_purged, ["ensemble2"])
    dataframe_ense3, te_report_ense3, tr_report_ense3 = to_dataframe(ensemble3_purged, ["ensemble3"])
    dataframe_ense4, te_report_ense4, tr_report_ense4 = to_dataframe(ensemble4_purged, ["ensemble4"])

    # join the dataframes
    all_data = pd.concat([dataframe_20, dataframe_40, dataframe_70, dataframe_80, dataframe_90, dataframe_ense1,
                          dataframe_ense2, dataframe_ense3, dataframe_ense4], axis=0)
    all_te_report = pd.concat([te_report_20, te_report_40, te_report_70, te_report_80, te_report_90,
                               te_report_ense1, te_report_ense2, te_report_ense3, te_report_ense4], axis=1)
    all_tr_report = pd.concat([tr_report_20, tr_report_40, tr_report_70, tr_report_80, tr_report_90,
                               tr_report_ense1, tr_report_ense2, tr_report_ense3, tr_report_ense4], axis=1)

    return all_data, all_te_report, all_tr_report


def writing(dataset1, te_report, tr_report, sheet_name, row=0, state=1):
    """Writes to excel"""
    if state == 1:
        state = "random"
    else:
        state = "ch2"
    if not path.exists(f"ensemble_scores/intramodel_knn_{state}.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename=f"ensemble_scores/intramodel_{state}.xlsx")

    book = load_workbook(f'ensemble_scores/intramodel_{state}.xlsx')
    writer = pd.ExcelWriter(f'ensemble_scores/intramodel_{state}.xlsx', engine='openpyxl')

    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    dataset1.to_excel(writer, sheet_name=f'{sheet_name}', startrow=row)
    tr_report.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(dataset1.index) + 3, )
    te_report.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(dataset1.index) + 3 + len(tr_report.index) + 3)

    writer.save()
    writer.close()


def run_esterase(name, state=1):
    """run all the script"""
    VHSE = pd.read_excel("sequences.xlsx", index_col=0, engine='openpyxl')
    Y = VHSE["label"].copy()
    X_knn_random = pd.read_excel("esterase_noerror.xlsx", index_col=0, sheet_name="random_30", engine='openpyxl')
    X_knn_ch = pd.read_excel("esterase_noerror.xlsx", index_col=0, sheet_name="ch2_20", engine='openpyxl')

    if X_knn_random.isnull().values.any():
        X_knn_random.dropna(axis=1, inplace=True)
        X_knn_random.drop(["go"], axis=1, inplace=True)
    if X_knn_ch.isnull().values.any():
        X_knn_ch.dropna(axis=1, inplace=True)
        X_knn_ch.drop(["go"], axis=1, inplace=True)

    # generating the predictions
    fitted_models = fit(state)
    for states in [20, 40, 70, 80, 90]:
        if state == 1:
            transformed_x_knn, test_x_knn, Y_train_knn, Y_test_knn, X_train_knn, X_test_knn = split_transform(X_knn_random, Y,
                                                                                                         states)
        else:
            transformed_x_knn, test_x_knn, Y_train_knn, Y_test_knn, X_train_knn, X_test_knn = split_transform(X_knn_ch, Y,
                                                                                                         states)
        test_pred, train_pred = predict(fitted_models, test_x_knn, transformed_x_knn)
        # get the scores
        all_data, all_te_report, all_tr_report = get_scores(Y_test_knn, Y_train_knn, test_pred, train_pred)

        writing(all_data, all_te_report, all_tr_report, f"{name}_{states}", state=state)


for state in range(1,3):
    run_esterase("knn", state)
