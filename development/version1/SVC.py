from sklearn.model_selection import train_test_split
from sklearn.svm import SVC
import pandas as pd
from sklearn.metrics import matthews_corrcoef, confusion_matrix
from sklearn.metrics import classification_report as class_re
from sklearn.model_selection import GridSearchCV, RandomizedSearchCV
from scipy.stats import uniform
from numpy import arange
import numpy
from collections import namedtuple
from sklearn.preprocessing import MinMaxScaler
from os import path
from openpyxl import load_workbook
from openpyxl import Workbook


# The support vectors machine classifier
def svc_classification(X_train, Y_train, X_test, state=20):
    """A function that applies grid and random search to tune model and also gives a prediction"""
    # Creating a score and parameters to search from

    scoring = {"f1": "f1_weighted"}

    grid_param2 = [{"kernel": ["linear"], "C": arange(0.01, 1.1, 0.3)}, {"kernel": ["linear"], "C": range(1, 10, 2)},
                   {"kernel": ["rbf"], "C": arange(0.01, 1.1, 0.3), "gamma": arange(0.01, 1.1, 0.3)},
                   {"kernel": ["rbf"], "C": arange(1, 10, 2), "gamma": arange(1, 10, 2)},
                   {"kernel": ["rbf"], "C": arange(1, 10, 2), "gamma": arange(0.01, 1.1, 0.3)},
                   {"kernel": ["rbf"], "C": arange(0.01, 1.1, 0.3), "gamma": range(1, 10, 2)}]

    random_param = [{"kernel": ["linear"], "C": uniform(1, 10)}, {"kernel": ["linear"], "C": uniform(0.01, 0.99)},
                    {"kernel": ["rbf"], "C": uniform(1, 10), "gamma": uniform(1, 10)},
                    {"kernel": ["rbf"], "C": uniform(0.01, 0.99), "gamma": uniform(0.01, 0.99)},
                    {"kernel": ["rbf"], "C": uniform(0.01, 0.99), "gamma": uniform(1, 10)},
                    {"kernel": ["rbf"], "C": uniform(0.01, 0.99), "gamma": uniform(1, 10)}]

    # Model setting

    svc_grid = GridSearchCV(SVC(class_weight="balanced"), grid_param2, scoring=scoring, refit="f1", cv=5)
    svc_random = RandomizedSearchCV(SVC(class_weight="balanced"), random_param, scoring=scoring, refit="f1", cv=5, random_state=state)

    # Model training
    fitted_grid = svc_grid.fit(X_train, Y_train)
    fitted_random = svc_random.fit(X_train, Y_train)

    # Model predictions
    y_random = fitted_random.best_estimator_.predict(X_test)
    y_grid = fitted_grid.best_estimator_.predict(X_test)

    grid_train_predicted = fitted_grid.best_estimator_.predict(X_train)
    random_train_predicted = fitted_random.best_estimator_.predict(X_train)

    return fitted_grid, fitted_random, y_grid, y_random, random_train_predicted, grid_train_predicted


def print_score(fitted_grid, fitted_random, Y_test, y_random, y_grid, random_train_predicted, grid_train_predicted,
                Y_train):
    """ The function prints the scores of the models and the prediction performance """
    target_names = ["class 0", "class 1"]
    # Model comparison
    grid_score = fitted_grid.best_score_
    grid_params = fitted_grid.best_params_
    random_score = fitted_random.best_score_
    random_params = fitted_random.best_params_

    # Training scores
    random_train_confusion = confusion_matrix(Y_train, random_train_predicted)
    grid_train_confusion = confusion_matrix(Y_train, grid_train_predicted)
    g_train_matthews = matthews_corrcoef(Y_train, grid_train_predicted)
    r_train_matthews = matthews_corrcoef(Y_train, random_train_predicted)
    random_tr_report = class_re(Y_train, random_train_predicted, target_names=target_names, output_dict=True)
    grid_tr_report = class_re(Y_train, grid_train_predicted, target_names=target_names, output_dict=True)
    # Test metrics
    random_confusion = confusion_matrix(Y_test, y_random)
    random_matthews = matthews_corrcoef(Y_test, y_random)
    random_te_report = class_re(Y_test, y_random, target_names=target_names, output_dict=True)

    grid_matthews = matthews_corrcoef(Y_test, y_grid)
    grid_confusion = confusion_matrix(Y_test, y_grid)
    grid_te_report = class_re(Y_test, y_grid, target_names=target_names, output_dict=True)

    return grid_score, grid_params, grid_confusion, random_tr_report, random_te_report, grid_te_report, grid_tr_report,\
           grid_matthews, random_score, random_params, random_confusion, random_matthews, \
           random_train_confusion, grid_train_confusion, r_train_matthews, g_train_matthews

def nested_cv(X, Y):
    """Performs something similar to a nested cross-validation"""

    metric_list = []
    model_list = []
    parameter_list = []
    random_state = [20, 40, 70, 80, 90]
    scaling = MinMaxScaler()
    esterase = ['EH51(22)', 'EH75(16)', 'EH46(23)', 'EH98(11)', 'EH49(23)']

    for states in random_state:
        X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.20, random_state=states, stratify=Y)
        X_train = X_train.loc[[x for x in X_train.index if x not in esterase]]
        X_test = X_test.loc[[x for x in X_test.index if x not in esterase]]
        Y_train = Y_train.loc[[x for x in Y_train.index if x not in esterase]]
        Y_test = Y_test.loc[[x for x in Y_test.index if x not in esterase]]

        transformed_x = scaling.fit_transform(X_train)
        transformed_x = pd.DataFrame(transformed_x)
        transformed_x.index = X_train.index
        transformed_x.columns = X_train.columns

        test_x = scaling.transform(X_test)
        test_x = pd.DataFrame(test_x)
        test_x.index = X_test.index
        test_x.columns = X_test.columns

        fitted_grid, fitted_random, y_grid, y_random, random_train_predicted, grid_train_predicted = svc_classification(
                transformed_x, Y_train, test_x)

        grid_score, grid_params, grid_confusion, random_tr_report, random_te_report, grid_te_report, grid_tr_report, \
        grid_matthews, random_score, random_params, random_confusion, random_matthews, \
        random_train_confusion, grid_train_confusion, r_train_matthews, g_train_matthews = print_score(
        fitted_grid, fitted_random, Y_test, y_random, y_grid, random_train_predicted, grid_train_predicted,Y_train)

        model_list.append([fitted_grid, fitted_random, y_grid, y_random])

        metric_list.append(
            [grid_score, grid_matthews, random_score,
             random_matthews, r_train_matthews, g_train_matthews])

        parameter_list.append([grid_params, grid_confusion, random_params, random_confusion,
                               random_train_confusion, grid_train_confusion, random_tr_report, random_te_report,
                               grid_te_report, grid_tr_report])

    return model_list, metric_list, parameter_list, random_state, scaling

def mean_nested(X, Y):
    """From the results of the nested_CV it computes the means of the different performance metrics"""
    model_list, metric_list, parameter_list, random_state, scaling = nested_cv(X, Y)
    score_record = namedtuple("scores", ["grid_score", "grid_matthews", "random_score", "random_matthews", "rt_matthews", "gt_matthews"])

    parameter_record = namedtuple("parameters", ["grid_params", "grid_confusion", "random_params", "random_confusion",
                                                 "random_train_confusion", "grid_train_confusion", "r_tr_report",
                                                 "r_te_report", "g_te_report", "g_tr_report"])

    model_record = namedtuple("models", ["fitted_grid", "fitted_random", "y_grid", "y_random"])

    array = numpy.array(metric_list)
    mean = numpy.mean(array, axis=0)

    named_parameters = [parameter_record(*z) for z in parameter_list]
    named_mean = score_record(*mean)
    named_records = [score_record(*y) for y in metric_list]
    named_models = [model_record(*d) for d in model_list]

    return named_mean, named_models, named_parameters, named_records, random_state, scaling

def unlisting(named_parameters, named_records):
    """ A function that separates all the scores in independent lists"""
    # Getting all scores random search
    r_mathew = [x.random_matthews for x in named_records]
    r_cv_score = [x.random_score for x in named_records]
    rt_mathew = [x.rt_matthews for x in named_records]
    # Getting all scores grid search
    g_mathew = [x.grid_matthews for x in named_records]
    g_cv_score = [x.grid_score for x in named_records]
    gt_mathew = [x.gt_matthews for x in named_records]

    # Hyperparameters grid search
    g_kernel = [y.grid_params["kernel"] for y in named_parameters]
    g_C = [y.grid_params["C"] for y in named_parameters]
    g_gamma = [y.grid_params["gamma"] if y.grid_params.get("gamma") else 0 for y in named_parameters]

    # Hyperparameters random search
    r_kernel = [y.random_params["kernel"] for y in named_parameters]
    r_C = [y.random_params["C"] for y in named_parameters]
    r_gamma = [y.random_params["gamma"] if y.random_params.get("gamma") else 0 for y in named_parameters]

    return r_mathew, g_mathew, g_kernel, g_C, g_gamma, r_kernel, r_C, r_gamma, r_cv_score, g_cv_score, rt_mathew, gt_mathew

def to_dataframe(named_parameters, named_records, random_state):
    matrix = namedtuple("confusion_matrix", ["true_n", "false_p", "false_n", "true_p"])

    r_mathew, g_mathew, g_kernel, g_C, g_gamma, r_kernel, r_C, r_gamma, r_cv_score, g_cv_score, \
    rt_mathew, gt_mathew = unlisting(named_parameters, named_records)

    # Taking the confusion matrix
    g_test_confusion = [matrix(*x.grid_confusion.ravel()) for x in named_parameters]
    r_test_confusion = [matrix(*x.random_confusion.ravel()) for x in named_parameters]

    g_training_confusion = [matrix(*x.grid_train_confusion.ravel()) for x in named_parameters]
    r_training_confusion = [matrix(*x.random_train_confusion.ravel())for x in named_parameters]

    # Otrher scores
    r_tr_report = [pd.DataFrame(x.r_tr_report).transpose() for x in named_parameters]
    r_te_report = [pd.DataFrame(x.r_te_report).transpose() for x in named_parameters]
    g_te_report = [pd.DataFrame(x.g_te_report).transpose() for x in named_parameters]
    g_tr_report = [pd.DataFrame(x.g_tr_report).transpose() for x in named_parameters]

    # Separating confusion matrix into individual elements
    g_test_true_n = [y.true_n for y in g_test_confusion]
    g_test_false_p = [y.false_p for y in g_test_confusion]
    g_test_false_n = [y.false_n for y in g_test_confusion]
    g_test_true_p = [y.true_p for y in g_test_confusion]

    g_training_true_n = [z.true_n for z in g_training_confusion]
    g_training_false_p = [z.false_p for z in g_training_confusion]
    g_training_false_n = [z.false_n for z in g_training_confusion]
    g_training_true_p = [z.true_p for z in g_training_confusion]

    r_test_true_n = [y.true_n for y in r_test_confusion]
    r_test_false_p = [y.false_p for y in r_test_confusion]
    r_test_false_n = [y.false_n for y in r_test_confusion]
    r_test_true_p = [y.true_p for y in r_test_confusion]

    r_training_true_n = [z.true_n for z in r_training_confusion]
    r_training_false_p = [z.false_p for z in r_training_confusion]
    r_training_false_n = [z.false_n for z in g_training_confusion]
    r_training_true_p = [z.true_p for z in g_training_confusion]


    r_dataframe = pd.DataFrame([random_state, r_kernel, r_C, r_gamma, r_test_true_n, r_test_true_p,
                    r_test_false_p, r_test_false_n, r_training_true_n, r_training_true_p, r_training_false_p,
                    r_training_false_n, r_mathew, rt_mathew, r_cv_score])

    g_dataframe = pd.DataFrame([random_state, g_kernel, g_C, g_gamma, g_test_true_n, g_test_true_p,
                g_test_false_p, g_test_false_n, g_training_true_n, g_training_true_p, g_training_false_p,
                g_training_false_n,g_mathew, gt_mathew, g_cv_score])

    r_dataframe = r_dataframe.transpose()
    r_dataframe.columns =["random state","kernel", "r_C", "r_gamma",
                    "test_tn", "test_tp", "test_fp","test_fn", "train_tn", "train_tp", "train_fp",
                    "train_fn", "Mathews","train_Mat", "CV_F1"]

    g_dataframe = g_dataframe.transpose()
    g_dataframe.columns = ["random state","kernel", "C", "gamma", "test_tn",
                "test_tp", "test_fp","test_fn", "train_tn","train_tp", "train_fp",
                "train_fn", "Mathews","train_Mat", "CV_F1"]

    return r_dataframe, g_dataframe,r_tr_report, r_te_report, g_te_report, g_tr_report

def view_setting ():
    """ Sets the console view of how many columns the console displays"""
    desired_width=320
    pd.set_option('display.width', desired_width)
    numpy.set_printoptions(linewidth=desired_width)
    pd.set_option('display.max_columns', 14)

def writing(dataset1, g_te_report, g_tr_report, datase2, r_tr_report, r_te_report, sheet_name, mode=1, row=0):
    if not path.exists(f"svc_no_error.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename=f"svc_no_error.xlsx")

    book = load_workbook(f"svc_no_error.xlsx")
    writer = pd.ExcelWriter(f"svc_no_error.xlsx", engine='openpyxl')

    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    dataset1.to_excel(writer, sheet_name=f'{sheet_name}', startrow=row)
    g_train = pd.concat(g_tr_report, axis=1)
    g_test = pd.concat(g_te_report, axis=1)
    g_train.to_excel(writer, sheet_name=f'{sheet_name}', startrow=row, startcol=len(dataset1.columns)+ 3 )
    g_test.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(g_test.index) + 3, startcol=len(dataset1.columns) + 3)

    datase2.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(dataset1.index) + row + 3)
    r_train = pd.concat(r_tr_report, axis=1)
    r_test = pd.concat(r_te_report, axis=1)
    r_train.to_excel(writer, sheet_name=f'{sheet_name}', startrow=(len(g_test.index) +3)*2, startcol=len(datase2.columns)+3)
    r_test.to_excel(writer, sheet_name=f'{sheet_name}', startrow=(len(g_test.index) + 3) * 3,
                     startcol=len(datase2.columns) + 3)
    writer.save()
    writer.close()

def run(sheet, mode=1):
    """ A function that runs nested_cv several times"""
    # reading the data
    VHSE = pd.read_excel("sequences.xlsx", index_col=0, sheet_name="VHSE", engine='openpyxl')
    # Y = VHSE["label7"].copy()

    if mode == 1 :
        Y = VHSE["label1"].copy()
        X = pd.read_excel("esterase_noerror.xlsx", index_col=0, sheet_name=f"{sheet}", engine='openpyxl')
    elif mode == 2:
        Y = VHSE["label6"].copy()
        X = pd.read_excel("../../esterase/post_filter/binary_40.xlsx", index_col=0, sheet_name=f"{sheet}")
    else:
        Y = VHSE["label7"].copy()
        X = pd.read_excel("../../esterase/post_filter/binary_35.xlsx", index_col=0, sheet_name=f"{sheet}")

    if X.isnull().values.any():
        X.dropna(axis=1, inplace=True)
        X.drop(["go"], axis=1, inplace=True)

    # creating name tuple
    results = namedtuple("results", ["named_mean", "model_list", "named_parameters", "named_records"])
    data = namedtuple("dataframe", ["r_dataframe", "g_dataframe"])

    # Trying the nested CV
    named_mean_split, model_list_split, named_parameters_split, named_records_split, random_state, scaling = mean_nested(X, Y)
    # Generates the dataframe
    r_dataframe, g_dataframe, r_tr_report, r_te_report, g_te_report, \
    g_tr_report = to_dataframe(named_parameters_split, named_records_split, random_state)

    results_list = results(*[named_mean_split, model_list_split, named_parameters_split, named_records_split])
    data_list = data(*[r_dataframe, g_dataframe])

    return results_list, data_list, scaling, r_tr_report, r_te_report, g_te_report, g_tr_report

def run_all(mode=1):
    """ A function that will perform the run fro every sheet"""
    if mode == 1:
        book = load_workbook("esterase_noerror.xlsx")
    elif mode == 2:
        book = load_workbook("../../esterase/post_filter/binary_40.xlsx")
    else:
        book = load_workbook("../../esterase/post_filter/binary_35.xlsx")
    # creating a list
    result_list = []
    data_list = []
    scaler_list = []
    for ws in book.worksheets:
        if ws.title == "ch2_20":
            results, data, scaler, r_tr_report, r_te_report, g_te_report, g_tr_report = run(ws.title, mode)
            writing(data.g_dataframe, g_te_report, g_tr_report, data.r_dataframe, r_tr_report, r_te_report, ws.title, mode)
            result_list.append(results)
            data_list.append(data)
            scaler_list.append(scaler)

    return result_list, data_list, scaler_list


result_list3, data_list3, scaler_list3 = run_all(1)

