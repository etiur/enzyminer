from sklearn.model_selection import train_test_split
from sklearn.linear_model import RidgeClassifier as ridge
import pandas as pd
from sklearn.metrics import matthews_corrcoef, confusion_matrix
from sklearn.model_selection import GridSearchCV
from numpy import arange
from os import path
import numpy
from collections import namedtuple
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import classification_report as class_re
from openpyxl import load_workbook
from openpyxl import Workbook

# The support vectors machine classifier
def ridge_classification(X_train, Y_train, X_test):
    """A function that applies grid and random search to tune model and also gives a prediction"""
    # Creating a score and parameters to search from

    scoring = {"f1": "f1_weighted"}

    grid_param2 = [{"alpha": arange(0.01, 0.1, 0.01)}, {"alpha": arange(0.1, 1, 0.01)}, {"alpha": arange(0.001, 0.01, 0.001)},
                   {"alpha": range(1, 25, 1)}, {"alpha": range(25, 50, 1)}]

    # Model setting
    svc_grid = GridSearchCV(ridge(random_state=0), grid_param2, scoring=scoring, refit="f1", cv=10)

    # Model training
    fitted_grid = svc_grid.fit(X_train, Y_train)

    # Model predictions
    y_grid = fitted_grid.best_estimator_.predict(X_test)
    grid_train_predicted = fitted_grid.best_estimator_.predict(X_train)

    return fitted_grid, y_grid, grid_train_predicted

def print_score(fitted_grid, Y_test, y_grid,grid_train_predicted, Y_train, X_test):
    """ The function prints the scores of the models and the prediction performance """
    score_tuple = namedtuple("scores", ["grid_score", "grid_params", "grid_confusion", "tr_report", "te_report",
                                        "train_mat", "grid_matthews", "grid_train_confusion", "grid_r2"])
    target_names = ["class 0", "class 1"]
    # Model comparison
    grid_score = fitted_grid.best_score_
    grid_params = fitted_grid.best_params_
    grid_r2 = fitted_grid.best_estimator_.score(X_test, Y_test)

    # Training scores
    grid_train_confusion = confusion_matrix(Y_train, grid_train_predicted)
    g_train_matthews = matthews_corrcoef(Y_train, grid_train_predicted)
    grid_tr_report = class_re(Y_train, grid_train_predicted, target_names=target_names, output_dict=True)
    # Test metrics
    grid_confusion = confusion_matrix(Y_test, y_grid)
    grid_matthews = matthews_corrcoef(Y_test, y_grid)
    grid_te_report = class_re(Y_test, y_grid, target_names=target_names, output_dict=True)


    all_scores = score_tuple(*[grid_score, grid_params, grid_confusion, grid_tr_report, grid_te_report, g_train_matthews,
           grid_matthews, grid_train_confusion, grid_r2])

    return all_scores

def nested_cv(X, Y):
    """Performs something similar to a nested cross-validation"""

    metric_list = []
    model_list = []
    parameter_list = []

    random_state = [20, 40, 70, 80, 90]
    scaling = MinMaxScaler()
    esterase = ['EH51(22)', 'EH75(16)', 'EH46(23)', 'EH98(11)', 'EH49(23)']

    for states in random_state:
        X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.20, random_state=states, stratify=Y)

        X_train = X_train.loc[[x for x in X_train.index if x not in esterase]]
        X_test = X_test.loc[[x for x in X_test.index if x not in esterase]]
        Y_train = Y_train.loc[[x for x in Y_train.index if x not in esterase]]
        Y_test = Y_test.loc[[x for x in Y_test.index if x not in esterase]]

        transformed_x = scaling.fit_transform(X_train)
        transformed_x = pd.DataFrame(transformed_x)
        transformed_x.index = X_train.index
        transformed_x.columns = X_train.columns

        test_x = scaling.transform(X_test)
        test_x = pd.DataFrame(test_x)
        test_x.index = X_test.index
        test_x.columns = X_test.columns

        fitted_grid, y_grid, grid_train_predicted = ridge_classification(transformed_x, Y_train, test_x)

        all_score = print_score(fitted_grid, Y_test, y_grid, grid_train_predicted, Y_train, test_x)

        model_list.append([fitted_grid, y_grid])

        metric_list.append([all_score.grid_score, all_score.grid_matthews, all_score.train_mat, all_score.grid_r2])

        parameter_list.append(
            [all_score.grid_params, all_score.grid_confusion, all_score.grid_train_confusion, all_score.tr_report,
             all_score.te_report])


    return model_list, metric_list, parameter_list, random_state

def mean_nested(X, Y):
    """From the results of the nested_CV it computes the means of the different performance metrics"""
    model_list, metric_list, parameter_list, random_state = nested_cv(X, Y)
    score_record = namedtuple("scores", ["grid_score", "grid_matthews", "train_mat", "grid_r2"])
    parameter_record = namedtuple("parameters",
                                  ["grid_params", "grid_confusion", "grid_train_confusion", "tr_report", "te_report"])

    model_record = namedtuple("models", ["fitted_grid", "y_grid"])

    array = numpy.array(metric_list)
    mean = numpy.mean(array, axis=0)

    named_parameters = [parameter_record(*z) for z in parameter_list]
    named_mean = score_record(*mean)
    named_records = [score_record(*y) for y in metric_list]
    named_models = [model_record(*d) for d in model_list]

    return named_mean, named_models, named_parameters, named_records, random_state

def unlisting(named_parameters, named_records):
    """ A function that separates all the scores in independent lists"""

    # Getting all scores grid search
    g_mathew = [x.grid_matthews for x in named_records]
    g_cv_score = [x.grid_score for x in named_records]
    g_r2 = [x.grid_r2 for x in named_records]
    gt_mathew = [x.train_mat for x in named_records]
    # Hyperparameters grid search
    g_alpha = [y.grid_params["alpha"] for y in named_parameters]

    return g_mathew, gt_mathew, g_alpha, g_cv_score, g_r2

def to_dataframe(named_parameters, named_records, random_state):
    matrix = namedtuple("confusion_matrix", ["true_n", "false_p", "false_n", "true_p"])

    g_mathew, gt_mathew, g_alpha, g_cv_score, g_r2 = unlisting(named_parameters, named_records)

    # Taking the confusion matrix
    g_test_confusion = [matrix(*x.grid_confusion.ravel()) for x in named_parameters]
    g_training_confusion = [matrix(*x.grid_train_confusion.ravel()) for x in named_parameters]
    g_te_report = [pd.DataFrame(x.te_report).transpose() for x in named_parameters]
    g_tr_report = [pd.DataFrame(x.tr_report).transpose() for x in named_parameters]

    # Separating confusion matrix into individual elements
    g_test_true_n = [y.true_n for y in g_test_confusion]
    g_test_false_p = [y.false_p for y in g_test_confusion]
    g_test_false_n = [y.false_n for y in g_test_confusion]
    g_test_true_p = [y.true_p for y in g_test_confusion]

    g_training_true_n = [z.true_n for z in g_training_confusion]
    g_training_false_p = [z.false_p for z in g_training_confusion]
    g_training_false_n = [z.false_n for z in g_training_confusion]
    g_training_true_p = [z.true_p for z in g_training_confusion]

    g_dataframe = pd.DataFrame([random_state ,g_alpha, g_r2, g_test_true_n, g_test_true_p,
                                g_test_false_p, g_test_false_n, g_training_true_n, g_training_true_p,
                                g_training_false_p, g_training_false_n, g_mathew, gt_mathew, g_cv_score])

    g_dataframe = g_dataframe.transpose()
    g_dataframe.columns = ["random state","alpha", "r2","test_tn",
                           "test_tp", "test_fp", "test_fn", "train_tn", "train_tp", "train_fp",
                           "train_fn", "Mathews","train_Mat", "CV_F1"]

    return g_dataframe, g_te_report, g_tr_report

def view_setting():
    """ Sets the console view of how many columns the console displays"""
    desired_width = 320
    pd.set_option('display.width', desired_width)
    numpy.set_printoptions(linewidth=desired_width)
    pd.set_option('display.max_columns', 14)

def writing(dataset1, g_te_report, g_tr_report, sheet_name, mode=1, row=0):
    """Writes to excel"""
    if not path.exists(f"ridge_no_error.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename=f"ridge_no_error.xlsx")

    book = load_workbook(f"ridge_no_error.xlsx")
    writer = pd.ExcelWriter(f"ridge_no_error.xlsx", engine='openpyxl')

    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    dataset1.to_excel(writer, sheet_name=f'{sheet_name}', startrow=row)
    g_train = pd.concat(g_tr_report, axis=1)
    g_test = pd.concat(g_te_report, axis=1)
    g_train.to_excel(writer, sheet_name=f'{sheet_name}', startrow=row, startcol=len(dataset1.columns) + 3)
    g_test.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(g_test.index) + 3,
                    startcol=len(dataset1.columns) + 3)

    writer.save()
    writer.close()

def run(sheet, mode=1):
    """ A function that runs nested_cv several times"""
    # reading the data
    VHSE = pd.read_excel("sequences.xlsx", index_col=0, sheet_name="VHSE", engine='openpyxl')
    # Y = VHSE["label7"].copy()

    if mode == 1 :
        Y = VHSE["label1"].copy()
        X = pd.read_excel("esterase_noerror.xlsx", index_col=0, sheet_name=f"{sheet}", engine='openpyxl')
    elif mode == 2:
        Y = VHSE["label6"].copy()
        X = pd.read_excel("../../esterase/post_filter/binary_40.xlsx", index_col=0, sheet_name=f"{sheet}")
    else:
        Y = VHSE["label7"].copy()
        X = pd.read_excel("../../esterase/post_filter/binary_35.xlsx", index_col=0, sheet_name=f"{sheet}")

    if X.isnull().values.any():
        X.dropna(axis=1, inplace=True)
        X.drop(["go"], axis=1, inplace=True)

    # creating name tuple
    results = namedtuple("results", ["named_mean_split", "model_list_split", "named_parameters_split", "named_records_split"])
    data = namedtuple("dataframe", ["g_dataframe", "grid_coefs", "grid_intercept"])

    # creating list of results
    named_mean, model_list, named_parameters, named_records, random_state = mean_nested(X, Y)
    g_dataframe, g_te_report, g_tr_report = to_dataframe(named_parameters, named_records, random_state)

    grid_intercept = [x.fitted_grid.best_estimator_.intercept_ for x in model_list]
    grid_coefs = [x.fitted_grid.best_estimator_.coef_ for x in model_list]

    results_list = results(*[named_mean, model_list, named_parameters, named_records])
    data_list = data(*[g_dataframe, grid_coefs, grid_intercept])


    return results_list, data_list, g_te_report, g_tr_report


def run_all(mode=1):
    """ A function that will perform the run fro every sheet"""
    if mode == 1:
        book = load_workbook("esterase_noerror.xlsx")
    elif mode == 2:
        book = load_workbook("../../esterase/post_filter/binary_40.xlsx")
    else:
        book = load_workbook("../../esterase/post_filter/binary_35.xlsx")
    # creating a list
    result_list = []
    data_list = []
    scaler_list = []
    for ws in book.worksheets:
        if ws.title == "ch2_20":
            results, data, g_te_report, g_tr_report = run(ws.title, mode)
            writing(data.g_dataframe, g_te_report, g_tr_report, ws.title, mode)
            result_list.append(results)
            data_list.append(data)

    return result_list, data_list

result_list3, data_list3 = run_all(1)
