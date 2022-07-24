from sklearn.neighbors import KNeighborsClassifier as KNN
from openpyxl import load_workbook
from openpyxl import Workbook
from sklearn.linear_model import RidgeClassifier as RIDGE
from sklearn.svm import SVC
import pandas as pd
from collections import namedtuple
import numpy as np
from os import path
from utils_generate import split_transform, print_score, to_dataframe
import joblib
from sklearn.neighbors import NeighborhoodComponentsAnalysis as NCA
from sklearn.pipeline import Pipeline


def vote(*args):
    """
    Hard voting for the ensembles

    Parameters
    ___________
    args: list[arrays]
        A list of prediction arrays
    """
    vote_ = []
    index = []
    if len(args) == 2:
        mean = np.mean(args, axis=0)
        for s, x in enumerate(mean):
            if x == 1 or x == 0:
                vote_.append(int(x))
            else:
                vote_.append(args[-1][s])
                index.append(s)  # keep the index of non unanimous predictions

    elif len(args) % 2 == 1:
        mean = np.mean(args, axis=0)
        for s, x in enumerate(mean):
            if x == 1 or x == 0:
                vote_.append(int(x))
            elif x > 0.5:
                vote_.append(1)
                index.append(s)
            else:
                vote_.append(0)
                index.append(s)

    else:
        mean = np.mean(args, axis=0)
        for s, x in enumerate(mean):
            if x == 1 or x == 0:
                vote_.append(int(x))
            elif x > 0.5:
                vote_.append(1)
                index.append(s)
            elif x < 0.5:
                vote_.append(0)
                index.append(s)
            else:
                vote_.append(args[-1][s])
                index.append(s)

    return vote_, index


def fit(states, type_="random"):
    """A function that trains the classifiers and ensembles them"""
    # the features and the labels
    VHSE = pd.read_excel("sequences.xlsx", index_col=0, sheet_name="dataset", engine='openpyxl')
    Y = VHSE["label"].copy()
    X_svc = pd.read_excel("esterase_noerror.xlsx", index_col=0, sheet_name="ch2_20", engine='openpyxl')
    X_knn = pd.read_excel("esterase_noerror.xlsx", index_col=0, sheet_name="random_30", engine='openpyxl')

    if X_svc.isnull().values.any():
        X_svc.dropna(axis=1, inplace=True)
        X_svc.drop(["go"], axis=1, inplace=True)

    if X_knn.isnull().values.any():
        X_knn.dropna(axis=1, inplace=True)
        X_knn.drop(["go"], axis=1, inplace=True)

    # named_tuples
    models = namedtuple("models", ["svc20", "svc80", "ridge20","ridge40","ridge80", "knn20", "knn80"])
    test = namedtuple("test_samples", ["x_svc", "x_knn", "y_svc", "y_knn", "x_test_svc", "x_test_knn"])
    train = namedtuple("train_samples", ["svc_x", "knn_x", "svc_y", "knn_y", "x_train_svc", "x_train_knn"])

    # split and train
    transformed_x_svc, test_x_svc, Y_train_svc, Y_test_svc, X_train_svc, X_test_svc = split_transform(X_svc, Y, states)
    transformed_x_knn, test_x_knn, Y_train_knn, Y_test_knn, X_train_knn, X_test_knn = split_transform(X_knn, Y, states)

    # the 3 algorithms
    svc_20 = SVC(C=0.31, kernel="rbf", gamma=0.91)
    svc_80 = SVC(C=3, kernel="rbf", gamma=0.31)
    if type_ == "random":
        knn_20 = KNN(n_neighbors=4, p=4, metric="minkowski", n_jobs=-1)
        knn_80 = KNN(n_neighbors=4, p=2, metric="minkowski", n_jobs=-1)
        knn_20.fit(transformed_x_knn, Y_train_knn)
        knn_80.fit(transformed_x_knn, Y_train_knn)
    else:
        knn_20 = KNN(n_neighbors=10, p=5, metric="minkowski", n_jobs=-1)
        knn_80 = KNN(n_neighbors=7, p=1, metric="canberra", n_jobs=-1)
        nca = NCA(random_state=20)
        knn_20 = Pipeline(steps=[("nca", nca), ("knn", knn_20)])
        knn_80 = Pipeline(steps=[("nca", nca), ("knn", knn_80)])
        knn_20.fit(transformed_x_svc, Y_train_svc)
        knn_80.fit(transformed_x_svc, Y_train_svc)
    ridge_20 = RIDGE(alpha=8, random_state=0)
    ridge_40 = RIDGE(alpha=3, random_state=0)
    ridge_80 = RIDGE(alpha=0.51, random_state=0)

    # fit the 3 algorithms
    svc_20.fit(transformed_x_svc, Y_train_svc)
    svc_80.fit(transformed_x_svc, Y_train_svc)
    ridge_20.fit(transformed_x_svc, Y_train_svc)
    ridge_40.fit(transformed_x_svc, Y_train_svc)
    ridge_80.fit(transformed_x_svc, Y_train_svc)


    # save in namedtuples
    fitted_models = models(*[svc_20, svc_80, ridge_20, ridge_40, ridge_80, knn_20, knn_80])
    test_sample = test(*[test_x_svc, test_x_knn, Y_test_svc, Y_test_knn, X_test_svc, X_test_knn])
    train_sample = train(*[transformed_x_svc, transformed_x_knn, Y_train_svc, Y_train_knn, X_train_svc, X_train_knn])

    return fitted_models, test_sample, train_sample


def predict(fitted, test_x_svc, test_x_knn, transformed_x_svc, transformed_x_knn, type_="random"):
    """Using fitted models it makes predictions"""
    # name_tuples
    predictions = namedtuple("predictions", ["svc20", "svc80", "ridge20","ridge40","ridge80", "knn20", "knn80"])

    # predict on X_test
    svc20_pred = fitted.svc20.predict(test_x_svc)
    svc80_pred = fitted.svc80.predict(test_x_svc)
    ridge20_pred = fitted.ridge20.predict(test_x_svc)
    ridge40_pred = fitted.ridge40.predict(test_x_svc)
    ridge80_pred = fitted.ridge80.predict(test_x_svc)

    if type_ == "random":
        knn20_pred = fitted.knn20.predict(test_x_knn)
        knn80_pred = fitted.knn80.predict(test_x_knn)
        train_knn20 = fitted.knn20.predict(transformed_x_knn)
        train_knn80 = fitted.knn80.predict(transformed_x_knn)
    else:
        knn20_pred = fitted.knn20.predict(test_x_svc)
        knn80_pred = fitted.knn80.predict(test_x_svc)
        train_knn20 = fitted.knn20.predict(transformed_x_svc)
        train_knn80 = fitted.knn80.predict(transformed_x_svc)
    # predict on X_train
    train_svc20 = fitted.svc20.predict(transformed_x_svc)
    train_svc80 = fitted.svc80.predict(transformed_x_svc)
    train_ridge20 = fitted.ridge20.predict(transformed_x_svc)
    train_ridge40 = fitted.ridge40.predict(transformed_x_svc)
    train_ridge80 = fitted.ridge80.predict(transformed_x_svc)

    test_pred = predictions(*[svc20_pred, svc80_pred, ridge20_pred, ridge40_pred, ridge80_pred, knn20_pred, knn80_pred])
    train_pred = predictions(*[train_svc20,train_svc80, train_ridge20, train_ridge40, train_ridge80, train_knn20,
                               train_knn80])

    return test_pred, train_pred


def predict_trained(test_x_svc, test_x_knn, transformed_x_svc, transformed_x_knn, models="models", type_="random"):
    """Using fitted models it makes predictions"""
    # name_tuples
    predictions = namedtuple("predictions", ["svc20", "svc80", "ridge20", "ridge40", "ridge80", "knn20", "knn80"])

    svc20 = joblib.load(f"{models}/svc_20.pkl")
    svc80 = joblib.load(f"{models}/svc_80.pkl")
    ridge20 = joblib.load(f"{models}/ridge_20.pkl")
    ridge40 = joblib.load(f"{models}/ridge_20.pkl")
    ridge80 = joblib.load(f"{models}/ridge_80.pkl")
    knn20 = joblib.load(f"{models}/knn_20.pkl")
    knn80 = joblib.load(f"{models}/knn_80.pkl")

    # predict on X_test
    svc20_pred = svc20.predict(test_x_svc)
    svc80_pred = svc80.predict(test_x_svc)
    ridge20_pred = ridge20.predict(test_x_svc)
    ridge40_pred = ridge40.predict(test_x_svc)
    ridge80_pred = ridge80.predict(test_x_svc)

    if type_ == "random":
        knn20_pred = knn20.predict(test_x_knn)
        knn80_pred = knn80.predict(test_x_knn)
        train_knn20 = knn20.predict(transformed_x_knn)
        train_knn80 = knn80.predict(transformed_x_knn)
    else:
        knn20_pred = knn20.predict(test_x_svc)
        knn80_pred = knn80.predict(test_x_svc)
        train_knn20 = knn20.predict(transformed_x_svc)
        train_knn80 = knn80.predict(transformed_x_svc)
    # predict on X_train
    train_svc20 = svc20.predict(transformed_x_svc)
    train_svc80 = svc80.predict(transformed_x_svc)
    train_ridge20 = ridge20.predict(transformed_x_svc)
    train_ridge40 = ridge40.predict(transformed_x_svc)
    train_ridge80 = ridge80.predict(transformed_x_svc)


    test_pred = predictions(*[svc20_pred, svc80_pred, ridge20_pred, ridge40_pred, ridge80_pred, knn20_pred, knn80_pred])
    train_pred = predictions(*[train_svc20, train_svc80, train_ridge20, train_ridge40, train_ridge80, train_knn20,
                               train_knn80])

    return test_pred, train_pred


def get_scores(Y_test_svc, Y_test_knn, Y_train_svc, Y_train_knn, test_pred, train_pred):
    """Converts the scores into dataframes"""
    # ensembles the predictions
    vote_1_test, index1_test = vote(test_pred.svc20, test_pred.ridge20)
    vote_1_train, index1_train = vote(train_pred.svc20, train_pred.ridge20)
    vote_2_test, index2_test = vote(test_pred.svc20, test_pred.ridge20, test_pred.knn20)
    vote_2_train, index2_train = vote(train_pred.svc20, train_pred.ridge20, train_pred.knn20)
    vote_3_test, index3_test = vote(test_pred.svc20, test_pred.svc80, test_pred.ridge20, test_pred.knn20)
    vote_3_train, index3_train = vote(train_pred.svc20,train_pred.svc80, train_pred.ridge20, train_pred.knn20)
    vote_4_test, index4_test = vote(test_pred.svc20, test_pred.svc80, test_pred.ridge20, test_pred.ridge40,
                                    test_pred.knn20)
    vote_4_train, index4_train = vote(train_pred.svc20,train_pred.svc80, train_pred.ridge20, train_pred.ridge40,
                                      train_pred.knn20)
    vote_5_test, index5_test = vote(test_pred.svc20, test_pred.svc80, test_pred.ridge20, test_pred.ridge40,
                                    test_pred.ridge80, test_pred.knn20)
    vote_5_train, index5_train = vote(train_pred.svc20,train_pred.svc80, train_pred.ridge20, train_pred.ridge40,
                                      train_pred.ridge80, train_pred.knn20)
    vote_6_test, index6_test = vote(test_pred.svc20, test_pred.svc80, test_pred.ridge20, test_pred.ridge40,
                                    test_pred.ridge80, test_pred.knn20, test_pred.knn80)
    vote_6_train, index6_train = vote(train_pred.svc20,train_pred.svc80, train_pred.ridge20, train_pred.ridge40,
                                      train_pred.ridge80, train_pred.knn20, train_pred.knn80)
    # generating the scores
    scores_svc = print_score(Y_test_svc, test_pred.svc20, train_pred.svc20, Y_train_svc)
    scores_ridge = print_score(Y_test_svc, test_pred.ridge20, train_pred.ridge20, Y_train_svc)
    knn_scores = print_score(Y_test_knn, test_pred.knn20, train_pred.knn20, Y_train_knn)
    ensemble1 = print_score(Y_test_knn, vote_1_test, vote_1_train, Y_train_knn, index1_test, index1_train, mode=1)
    ensemble2 = print_score(Y_test_svc, vote_2_test, vote_2_train, Y_train_svc, index2_test, index2_train, mode=1)
    ensemble3 = print_score(Y_test_svc, vote_3_test, vote_3_train, Y_train_svc, index3_test, index3_train, mode=1)
    ensemble4 = print_score(Y_test_knn, vote_4_test, vote_4_train, Y_train_knn, index4_test, index4_train, mode=1)
    ensemble5 = print_score(Y_test_svc, vote_5_test, vote_5_train, Y_train_svc, index5_test, index5_train, mode=1)
    ensemble6 = print_score(Y_test_knn, vote_6_test, vote_6_train, Y_train_knn, index6_test, index6_train, mode=1)

    # put all the sores into dataframe
    dataframe_svc, te_report_svc, tr_report_svc = to_dataframe(scores_svc, ["svc"])
    dataframe_ridge, te_report_ridge, tr_report_ridge = to_dataframe(scores_ridge, ["ridge"])
    dataframe_knn, te_report_knn, tr_report_knn = to_dataframe(knn_scores, ["knn"])
    dataframe_ense1, te_report_ense1, tr_report_ense1 = to_dataframe(ensemble1, ["ensemble1"])
    dataframe_ense2, te_report_ense2, tr_report_ense2 = to_dataframe(ensemble2, ["ensemble2"])
    dataframe_ense3, te_report_ense3, tr_report_ense3 = to_dataframe(ensemble3, ["ensemble3"])
    dataframe_ense4, te_report_ense4, tr_report_ense4 = to_dataframe(ensemble4, ["ensemble4"])
    dataframe_ense5, te_report_ense5, tr_report_ense5 = to_dataframe(ensemble5, ["ensemble5"])
    dataframe_ense6, te_report_ense6, tr_report_ense6 = to_dataframe(ensemble6, ["ensemble6"])
    # join the dataframes
    all_data = pd.concat([dataframe_svc, dataframe_ridge, dataframe_knn, dataframe_ense1, dataframe_ense2,
                          dataframe_ense3, dataframe_ense4, dataframe_ense5, dataframe_ense6], axis=0)
    all_te_report = pd.concat([te_report_svc, te_report_ridge, te_report_knn, te_report_ense1, te_report_ense2,
                               te_report_ense3, te_report_ense4, te_report_ense5, te_report_ense6], axis=1)
    all_tr_report = pd.concat([tr_report_svc, tr_report_ridge, tr_report_knn, tr_report_ense1, tr_report_ense2,
                               tr_report_ense3, tr_report_ense4, tr_report_ense5, tr_report_ense6], axis=1)

    return all_data, all_te_report, all_tr_report


def writing(dataset1, te_report, tr_report, sheet_name, row=0, type_="random"):
    """Writes to excel"""
    if not path.exists(f"ensemble_scores/ensemble_noerror_{type_}.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename=f"ensemble_scores/ensemble_noerror_{type_}.xlsx")

    book = load_workbook(f'ensemble_scores/ensemble_noerror_{type_}.xlsx')
    writer = pd.ExcelWriter(f'ensemble_scores/ensemble_noerror_{type_}.xlsx', engine='openpyxl')

    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    dataset1.to_excel(writer, sheet_name=f'{sheet_name}', startrow=row)
    tr_report.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(dataset1.index)+3, )
    te_report.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(dataset1.index)+3+len(tr_report.index)+3)

    writer.save()
    writer.close()


def run_esterase(name, type_="random"):
    """run all the script"""
    VHSE = pd.read_excel("sequences.xlsx", index_col=0, sheet_name="dataset", engine='openpyxl')
    Y = VHSE["label"].copy()
    X_svc = pd.read_excel("esterase_noerror.xlsx", index_col=0, sheet_name="ch2_20", engine='openpyxl')
    X_knn = pd.read_excel("esterase_noerror.xlsx", index_col=0, sheet_name="random_30", engine='openpyxl')

    if X_svc.isnull().values.any():
        X_svc.dropna(axis=1, inplace=True)
        X_svc.drop(["go"], axis=1, inplace=True)

    if X_knn.isnull().values.any():
        X_knn.dropna(axis=1, inplace=True)
        X_knn.drop(["go"], axis=1, inplace=True)

    # generating the predictions
    for states in [20, 40, 70, 80, 90]:
        fitted_models, test_sample, train_sample = fit(states, type_)
        transformed_x_svc, test_x_svc, Y_train_svc, Y_test_svc, X_train_svc, X_test_svc = split_transform(X_svc, Y, states)
        transformed_x_knn, test_x_knn, Y_train_knn, Y_test_knn, X_train_knn, X_test_knn = split_transform(X_knn, Y, states)

        #test_pred, train_pred = predict_trained(test_x_svc, test_x_knn, transformed_x_svc, transformed_x_knn, type_=type_)

        test_pred, train_pred = predict(fitted_models, test_x_svc, test_x_knn, transformed_x_svc, transformed_x_knn, type_=type_)
    # get the scores
        all_data, all_te_report, all_tr_report = get_scores(Y_test_svc, Y_test_knn, Y_train_svc,
                                                        Y_train_knn, test_pred, train_pred)

        writing(all_data, all_te_report, all_tr_report, f"{name}_{states}", type_=type_)

for type_ in ["random", "ch2"]:
    run_esterase("esterase", type_)



